import os, io, base64, requests
from datetime import datetime
from flask import request, send_file, Blueprint
from flask_jwt_extended import get_jwt_identity
from sqlalchemy import func, case, text, extract
from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from PIL import Image
from io import BytesIO
import matplotlib
matplotlib.use("Agg")
from matplotlib import pyplot as plt
from models.schemes import Usuario, Empresa, Preferencias_empresa, Oferta_laboral, Job_Application, RendimientoEmpleado, Licencia, Rol, UsuarioRol, Oferta_analista
from auth.decorators import role_required
from .manager import manager_bp
from models.extensions import db
from collections import defaultdict


reportes_bp = Blueprint('reportes_bp', __name__)

BASE_DIR = os.getcwd()
TEMP_DIR = os.path.join(os.path.dirname(__file__), '..', 'temp')     # <-- AGREGADO POR JOA PARA PROBAR LA PLANTILLA, EVALUAR SI LES SIRVE 
os.makedirs(TEMP_DIR, exist_ok=True)


@reportes_bp.route("/reportes-reclutamiento", methods=["GET"])
@role_required(["manager", "reclutador"])
def reporte_reclutamiento():
    formato = request.args.get("formato", "pdf")
    ids_str = request.args.get("ids")
    ids_ofertas = list(map(int, ids_str.split(","))) if ids_str else None

    id_manager = get_jwt_identity()
    manager = Usuario.query.get(id_manager)
    preferencia = Preferencias_empresa.query.get(manager.id_empresa)
    empresa = Empresa.query.get(manager.id_empresa)

    query = (
        db.session.query(
            Oferta_laboral.nombre.label("oferta"),
            func.count(Job_Application.id).label("postulaciones"),
            func.sum(case((Job_Application.is_apto == True, 1), else_=0)).label("contratados"),
            func.avg(func.timestampdiff(
                text("DAY"),
                Oferta_laboral.fecha_publicacion,
                Job_Application.fecha_postulacion
            )).label("tiempo_promedio")
        )
        .outerjoin(Job_Application, Job_Application.id_oferta == Oferta_laboral.id)
        .filter(Oferta_laboral.id_empresa == manager.id_empresa)
    )

    if ids_ofertas:
        query = query.filter(Oferta_laboral.id.in_(ids_ofertas))
        
    query = query.group_by(Oferta_laboral.id, Oferta_laboral.nombre)
        
    resultados = query.group_by(Oferta_laboral.id).all()
    
    if not resultados:
        return {"error": "No se encontraron ofertas con los IDs proporcionados."}, 404

    datos = [
        {
            "oferta": row.oferta,
            "postulaciones": row.postulaciones or 0,
            "contratados": row.contratados or 0,
            "tasa_conversion": round(((row.contratados or 0) / (row.postulaciones or 1)) * 100, 1) if row.postulaciones else 0,
            "tiempo_promedio": round(row.tiempo_promedio or 0, 1)
        }
        for row in resultados
    ]

    color_secundario = preferencia.color_principal if preferencia and preferencia.color_principal else "2E86C1"
    color_hex = color_secundario if color_secundario.startswith("#") else f"#{color_secundario}"
    
    if formato == "pdf":
        env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), '..', 'templates')))
        template = env.get_template("reporte_reclutamiento_profesional.html")

        
        def generar_grafico(datos, campo, titulo, nombre_archivo):
            etiquetas = [d["oferta"] for d in datos]
            valores = [d[campo] for d in datos]

            fig, ax = plt.subplots()
            ax.bar(etiquetas, valores, color=color_hex)
            ax.set_title(titulo)
            ax.set_ylabel(campo.replace("_", " ").capitalize())
            ax.set_xticks(range(len(etiquetas)))
            ax.set_xticklabels(etiquetas, rotation=45, ha='right')

            img = BytesIO()
            plt.tight_layout()
            plt.savefig(img, format="png", dpi=150)
            plt.close(fig)
            img.seek(0)
            return base64.b64encode(img.getvalue()).decode("utf-8")
        
        grafico_postulaciones = generar_grafico(datos, "postulaciones", "Postulaciones por Oferta", "postulaciones.png")
        grafico_conversion = generar_grafico(datos, "tasa_conversion", "Tasa de Conversión", "conversion.png")
        grafico_tiempo = generar_grafico(datos, "tiempo_promedio", "Tiempo Promedio de Postulación", "tiempo.png")

        now = datetime.now()

        html_out = template.render(
            empresa=empresa.nombre,
            logo_url=preferencia.logo_url if preferencia and preferencia.logo_url else None,
            color=color_hex,
            datos=datos,
            now=now,
            grafico_postulaciones=grafico_postulaciones,
            grafico_conversion=grafico_conversion,
            grafico_tiempo=grafico_tiempo
        )
        
        TEMP_DIR = os.path.join(os.path.dirname(__file__), '..', 'temp')
        os.makedirs(TEMP_DIR, exist_ok=True)
        nombre_archivo = f"informe_reclutamiento_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta_archivo = os.path.abspath(os.path.join(TEMP_DIR, nombre_archivo))
        HTML(string=html_out).write_pdf(ruta_archivo)
        return send_file(ruta_archivo, as_attachment=True, download_name=nombre_archivo)
    
    if formato == "excel":

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Reclutamiento"

        if preferencia and preferencia.logo_url:
            try:
                response = requests.get(preferencia.logo_url)
                if response.status_code == 200:
                    ruta_logo = os.path.join("temp", "logo_empresa.png")
                    with open(ruta_logo, "wb") as f:
                        f.write(response.content)
                    img = ExcelImage(ruta_logo)
                    img.width, img.height = 100, 50
                    ws.add_image(img, "A1")
            except:
                pass

        ws.merge_cells("A4:E4")
        ws["A4"] = f"Informe de Reclutamiento - {empresa.nombre}"
        ws["A4"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A4"].alignment = Alignment(horizontal="center")
        ws["A4"].fill = PatternFill(start_color=color_hex.replace("#", ""), end_color=color_hex.replace("#", ""), fill_type="solid")

        headers = ["Oferta", "Postulaciones", "Contratados", "Tasa de Conversión (%)", "Tiempo Promedio (días)"]
        ws.append(headers)
        header_row = ws.max_row
        for col in range(1, 6):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color_hex.replace("#", ""), end_color=color_hex.replace("#", ""), fill_type="solid")

        for fila in datos:
            ws.append([
                fila["oferta"],
                fila["postulaciones"],
                fila["contratados"],
                fila["tasa_conversion"],
                fila["tiempo_promedio"]
            ])

        last_row = ws.max_row

        chart1 = BarChart()
        chart1.title = "Postulaciones por Oferta"
        chart1.y_axis.title = "Postulaciones"
        chart1.x_axis.title = "Ofertas"
        data = Reference(ws, min_col=2, min_row=header_row, max_row=last_row)
        categories = Reference(ws, min_col=1, min_row=header_row+1, max_row=last_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(categories)
        chart1.height = 7
        chart1.width = 16
        ws.add_chart(chart1, "G5")

        chart2 = BarChart()
        chart2.title = "Postulaciones vs Contratados"
        data = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=last_row)
        categories = Reference(ws, min_col=1, min_row=header_row+1, max_row=last_row)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(categories)
        chart2.height = 7
        chart2.width = 16
        ws.add_chart(chart2, "G22")

        chart3 = LineChart()
        chart3.title = "Tiempo Promedio de Postulación"
        chart3.y_axis.title = "Días"
        chart3.x_axis.title = "Ofertas"
        data = Reference(ws, min_col=5, min_row=header_row, max_row=last_row)
        categories = Reference(ws, min_col=1, min_row=header_row+1, max_row=last_row)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(categories)
        chart3.height = 7
        chart3.width = 16
        ws.add_chart(chart3, "G39")


        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = max_length + 2
            ws.column_dimensions[col_letter].width = adjusted_width

        TEMP_DIR = os.path.join(os.path.dirname(__file__), '..', 'temp')
        os.makedirs(TEMP_DIR, exist_ok=True)
        archivo_excel = f"informe_reclutamiento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_excel = os.path.abspath(os.path.join(TEMP_DIR, archivo_excel))
        wb.save(ruta_excel)
        return send_file(ruta_excel, as_attachment=True, download_name=archivo_excel)

    return {"error": "Formato no soportado"}, 400


def generar_grafico_promedio_puesto_base64(data):
    puestos = [p["puesto"] for p in data]
    promedios = [p["promedio"] for p in data]

    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.barh(puestos, promedios, color="#117A65")
    ax.set_title("Promedio de Rendimiento Futuro por Puesto", fontsize=14)
    ax.set_xlabel("Promedio", fontsize=12)
    ax.set_xlim(0, 10)

    for bar in bars:
        width = bar.get_width()
        ax.annotate(f'{width:.1f}', xy=(width, bar.get_y() + bar.get_height() / 2),
                    xytext=(5, 0), textcoords="offset points", ha='left', va='center', fontsize=10)

    plt.tight_layout()
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")

def generar_grafico_base64(data):
    nombres = [f["nombre"] for f in data]
    valores = [f["rendimiento_futuro"] for f in data]

    fig, ax = plt.subplots(figsize=(10, 4))
    bars = ax.bar(nombres, valores, color="#2E86C1")
    ax.set_title("Top Rendimiento Futuro")
    ax.set_ylabel("Predicción")
    ax.set_ylim(0, 10)
    ax.set_xticklabels(nombres, rotation=45, ha='right')

    for bar in bars:
        height = bar.get_height()
        ax.annotate(f'{height:.1f}', xy=(bar.get_x() + bar.get_width() / 2, height),
                    xytext=(0, 3), textcoords="offset points", ha='center', fontsize=10)

    plt.tight_layout()
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


def grafico_ausencias_base64(data):
    meses = [f["mes"] for f in data]
    dias = [f["total_dias"] for f in data]

    fig, ax = plt.subplots(figsize=(10, 4))

    if not dias:
        ax.text(0.5, 0.5, 'Sin datos de ausencias', ha='center', va='center', fontsize=14)
        ax.axis('off')
    else:
        bars = ax.bar(meses, dias, color="#C0392B")
        ax.set_title("Días de Ausencia por Mes", fontsize=14)
        ax.set_ylabel("Días", fontsize=12)
        ax.set_xlabel("Mes", fontsize=12)
        ax.set_ylim(0, max(dias)+5)

        for bar in bars:
            height = bar.get_height()
            ax.annotate(f'{height}', xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3), textcoords="offset points", ha='center', fontsize=10)

    plt.tight_layout()
    buffer = io.BytesIO()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")



    # env = Environment(loader=FileSystemLoader("templates"))
    # template = env.get_template("reporte_rendimiento.html")

    # if formato == "pdf":
    #     html_out = template.render(
    #         empresa=empresa.nombre,
    #         logo_url=preferencia.logo_url if preferencia and preferencia.logo_url else None,
    #         color=preferencia.color_secundario or "#2E86C1",
    #         ranking_futuro=ranking_dict,
    #         grafico_base64=grafico_base64,
    #         promedios_por_puesto=promedios_por_puesto,
    #         grafico_puesto_base64=grafico_puesto_base64
    #     )
    
    #     nombre_archivo = f"desempeno_futuro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    #     ruta_archivo = os.path.join("temp", nombre_archivo)
    #     os.makedirs("temp", exist_ok=True)
    #     HTML(string=html_out).write_pdf(ruta_archivo)

    #     return send_file(ruta_archivo, as_attachment=True, download_name=nombre_archivo)



@reportes_bp.route("/reportes-desempeno", methods=["GET"])
@role_required(["manager", "reclutador"])
def reporte_desempeno():
    formato = request.args.get("formato", "pdf")

    id_manager = get_jwt_identity()
    manager = Usuario.query.get(id_manager)
    preferencia = Preferencias_empresa.query.get(manager.id_empresa)
    empresa = Empresa.query.get(manager.id_empresa)

    color = preferencia.color_principal[1:] if preferencia and preferencia.color_principal else "2E86C1"

    ranking_futuro = (
        db.session.query(
            Usuario.nombre,
            Usuario.apellido,
            Usuario.puesto_trabajo,
            Usuario.username,
            RendimientoEmpleado.rendimiento_futuro_predicho.label("prediccion"),
            RendimientoEmpleado.clasificacion_rendimiento,
            Rol.nombre.label("rol")
        )
        .join(Usuario, Usuario.id == RendimientoEmpleado.id_usuario)
        .join(UsuarioRol, Usuario.id == UsuarioRol.id_usuario)
        .join(Rol, UsuarioRol.id_rol == Rol.id)
        .filter(Usuario.id_empresa == manager.id_empresa)
        .filter(RendimientoEmpleado.rendimiento_futuro_predicho != None)
        .order_by(RendimientoEmpleado.rendimiento_futuro_predicho.desc())
        .limit(10)
        .all()
    )
    
    def clasificar_rendimiento(valor):
        if valor is None:
            return "Sin datos"
        if valor >= 8:
            return "Alto"
        elif valor >= 5:
            return "Medio"
        else:
            return "Bajo"

    ranking_dict = [
        {
            "nombre": r.nombre,
            "apellido": r.apellido,
            "puesto": r.puesto_trabajo or "Analista",
            "rol": r.rol,
            "username": r.username,
            "rendimiento_futuro": round(r.prediccion, 2),
            "clasificacion_rendimiento": r.clasificacion_rendimiento or clasificar_rendimiento(r.prediccion)
        }
        for r in ranking_futuro
    ]

    grafico_base64 = generar_grafico_base64(ranking_dict)

    agrupado_por_puesto = (
        db.session.query(
            Usuario.puesto_trabajo,
            func.avg(RendimientoEmpleado.rendimiento_futuro_predicho).label("promedio")
        )
        .join(RendimientoEmpleado, Usuario.id == RendimientoEmpleado.id_usuario)
        .filter(Usuario.id_empresa == empresa.id)
        .filter(RendimientoEmpleado.rendimiento_futuro_predicho != None)
        .group_by(Usuario.puesto_trabajo)
        .order_by(func.avg(RendimientoEmpleado.rendimiento_futuro_predicho).desc())
        .all()
    )

    promedios_por_puesto = [
        {
            "puesto": row.puesto_trabajo if row.puesto_trabajo else "Analista",
            "promedio": round(row.promedio, 2)
        }
        for row in agrupado_por_puesto
    ]

    grafico_puesto_base64 = generar_grafico_promedio_puesto_base64(promedios_por_puesto)

    if formato == "pdf":
        env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), '..', 'templates')))
        template = env.get_template("reporte_rendimiento.html")
        html_out = template.render(
            empresa=empresa.nombre,
            logo_url=preferencia.logo_url if preferencia and preferencia.logo_url else None,
            color=preferencia.color_principal if preferencia and preferencia.color_principal else "#2E86C1",            
            ranking_futuro=ranking_dict,
            grafico_base64=grafico_base64,
            promedios_por_puesto=promedios_por_puesto,
            grafico_puesto_base64=grafico_puesto_base64,
            now=datetime.now()
        )
        nombre_archivo = f"desempeno_futuro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta_archivo = os.path.abspath(os.path.join(TEMP_DIR, nombre_archivo))
        HTML(string=html_out).write_pdf(ruta_archivo)
        return send_file(ruta_archivo, as_attachment=True, download_name=nombre_archivo)

    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Predicción Rendimiento"

        color = preferencia.color_principal[1:] if preferencia and preferencia.color_principal else "2E86C1"

        # Título y logo
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Reporte Predicción de Rendimiento - {empresa.nombre}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        if preferencia and preferencia.logo_url:
            try:
                response = requests.get(preferencia.logo_url)
                if response.status_code == 200:
                    logo_path = os.path.join(TEMP_DIR, "logo_empresa.png")
                    with open(logo_path, "wb") as f:
                        f.write(response.content)
                    logo_img = ExcelImage(logo_path)
                    logo_img.width = 120
                    logo_img.height = 60
                    ws.add_image(logo_img, "E1")
            except Exception as e:
                print("No se pudo cargar el logo:", e)

        # Tabla Ranking
        # Tabla Ranking
        ws.append(["", "", "", "", "", ""])
        ws.append(["#", "Nombre", "Apellido", "Puesto", "Rol", "Usuario", "Rendimiento Futuro", "Clasificación"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for i, fila in enumerate(ranking_dict, start=1):
            ws.append([
                i,
                fila["nombre"],
                fila["apellido"],
                fila["puesto"] or "Analista",
                fila["rol"],
                fila["username"],
                fila["rendimiento_futuro"],
                fila["clasificacion_rendimiento"]
            ])

        # Tabla Promedios por Puesto
        ws.append(["", "", "", ""])
        ws.append(["Puesto", "Promedio de Rendimiento Futuro"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for fila in promedios_por_puesto:
            ws.append([fila["puesto"], fila["promedio"]])

        # Guardar gráficos dinámicos
        ruta_grafico1 = os.path.join(TEMP_DIR, "grafico_ranking.png")
        ruta_grafico2 = os.path.join(TEMP_DIR, "grafico_puesto.png")
        with open(ruta_grafico1, "wb") as f:
            f.write(base64.b64decode(grafico_base64))
        with open(ruta_grafico2, "wb") as f:
            f.write(base64.b64decode(grafico_puesto_base64))

        img1 = ExcelImage(ruta_grafico1)
        img2 = ExcelImage(ruta_grafico2)
        img1.width = img2.width = 600
        img1.height = img2.height = 300
        ws.add_image(img1, f"A{ws.max_row + 2}")
        ws.add_image(img2, f"A{ws.max_row + 20}")

        # Insertar imágenes externas del frontend
        imagenes_custom = [
            "cantidad_analistas.png",
            "distribucion_analistas.png",
            "promedios_analistas.png",
            "rendimiento_analistas.png"
        ]
        imagenes_path = os.path.join(os.getcwd(), "imagenes_reportes")
        fila_inicio = ws.max_row + 2
        columna_inicio = "G"

        for i, nombre in enumerate(imagenes_custom):
            ruta_imagen = os.path.join(imagenes_path, nombre)
            if os.path.exists(ruta_imagen):
                try:
                    img = ExcelImage(ruta_imagen)
                    img.width = 600
                    img.height = 300
                    celda = f"{columna_inicio}{fila_inicio + i * 18}"
                    ws.add_image(img, celda)
                except Exception as e:
                    print(f"Error insertando imagen {nombre}:", e)

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        archivo_excel = f"desempeno_futuro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_excel = os.path.join(TEMP_DIR, archivo_excel)
        wb.save(ruta_excel)

        return send_file(ruta_excel, as_attachment=True, download_name=archivo_excel)

    return {"error": "Formato no soportado"}, 400



@reportes_bp.route("/reportes-desempeno-analista", methods=["GET"])
@role_required(["reclutador"])
def reporte_desempeno_analista():
    formato = request.args.get("formato", "pdf")

    id_manager = get_jwt_identity()
    manager = Usuario.query.get(id_manager)
    preferencia = Preferencias_empresa.query.get(manager.id_empresa)
    empresa = Empresa.query.get(manager.id_empresa)

    color = preferencia.color_principal[1:] if preferencia and preferencia.color_principal else "2E86C1"
    
    ranking_futuro = (
        db.session.query(
            Usuario.nombre,
            Usuario.apellido,
            Usuario.puesto_trabajo,
            Usuario.username,
            RendimientoEmpleado.rendimiento_futuro_predicho.label("prediccion"),
            RendimientoEmpleado.clasificacion_rendimiento,
            Rol.nombre.label("rol")
        )
        .join(Usuario, Usuario.id == RendimientoEmpleado.id_usuario)
        .join(UsuarioRol, Usuario.id == UsuarioRol.id_usuario)
        .join(Rol, UsuarioRol.id_rol == Rol.id)
        .filter(Usuario.id_empresa == manager.id_empresa)
        .filter(Rol.slug == "empleado")
        .filter(RendimientoEmpleado.rendimiento_futuro_predicho != None)
        .order_by(RendimientoEmpleado.rendimiento_futuro_predicho.desc())
        .limit(10)
        .all()
    )
    
    def clasificar_rendimiento(valor):
        if valor is None:
            return "Sin datos"
        if valor >= 8:
            return "Alto"
        elif valor >= 5:
            return "Medio"
        else:
            return "Bajo"

    ranking_dict = [
        {
            "nombre": r.nombre,
            "apellido": r.apellido,
            "puesto": r.puesto_trabajo or "Analista",
            "rol": r.rol,
            "username": r.username,
            "rendimiento_futuro": round(r.prediccion, 2),
            "clasificacion_rendimiento": r.clasificacion_rendimiento or clasificar_rendimiento(r.prediccion)
        }
        for r in ranking_futuro
    ]
    
    grafico_base64 = generar_grafico_base64(ranking_dict)

    agrupado_por_puesto = (
        db.session.query(
            Usuario.puesto_trabajo,
            func.avg(RendimientoEmpleado.rendimiento_futuro_predicho).label("promedio")
        )
        .join(RendimientoEmpleado, Usuario.id == RendimientoEmpleado.id_usuario)
        .join(UsuarioRol, Usuario.id == UsuarioRol.id_usuario)
        .join(Rol, Rol.id == UsuarioRol.id_rol)
        .filter(Usuario.id_empresa == empresa.id)
        .filter(Rol.slug == "empleado")
        .filter(RendimientoEmpleado.rendimiento_futuro_predicho != None)
        .group_by(Usuario.puesto_trabajo)
        .order_by(func.avg(RendimientoEmpleado.rendimiento_futuro_predicho).desc())
        .all()
    )

    promedios_por_puesto = [
        {
            "puesto": row.puesto_trabajo if row.puesto_trabajo else "Analista",
            "promedio": round(row.promedio, 2)
        }
        for row in agrupado_por_puesto
    ]

    grafico_puesto_base64 = generar_grafico_promedio_puesto_base64(promedios_por_puesto)

    if formato == "pdf":
        env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), '..', 'templates')))
        template = env.get_template("reporte_rendimiento.html")
        html_out = template.render(
            empresa=empresa.nombre,
            logo_url=preferencia.logo_url if preferencia and preferencia.logo_url else None,
            color=preferencia.color_principal if preferencia and preferencia.color_principal else "#2E86C1",            
            ranking_futuro=ranking_dict,
            grafico_base64=grafico_base64,
            promedios_por_puesto=promedios_por_puesto,
            grafico_puesto_base64=grafico_puesto_base64,
            now=datetime.now()
        )
        nombre_archivo = f"desempeno_futuro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta_archivo = os.path.abspath(os.path.join(TEMP_DIR, nombre_archivo))
        HTML(string=html_out).write_pdf(ruta_archivo)
        return send_file(ruta_archivo, as_attachment=True, download_name=nombre_archivo)

    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Predicción Rendimiento"

        color = preferencia.color_principal[1:] if preferencia and preferencia.color_principal else "2E86C1"

        # Título y logo
        ws.merge_cells("A1:D1")
        ws["A1"] = f"Reporte Predicción de Rendimiento - {empresa.nombre}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        if preferencia and preferencia.logo_url:
            try:
                response = requests.get(preferencia.logo_url)
                if response.status_code == 200:
                    logo_path = os.path.join(TEMP_DIR, "logo_empresa.png")
                    with open(logo_path, "wb") as f:
                        f.write(response.content)
                    logo_img = ExcelImage(logo_path)
                    logo_img.width = 120
                    logo_img.height = 60
                    ws.add_image(logo_img, "E1")
            except Exception as e:
                print("No se pudo cargar el logo:", e)

        # Tabla Ranking
        ws.append(["", "", "", "", "", ""])
        ws.append(["#", "Nombre", "Apellido", "Puesto", "Rol", "Usuario", "Rendimiento Futuro", "Clasificación"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for i, fila in enumerate(ranking_dict, start=1):
            ws.append([
                i,
                fila["nombre"],
                fila["apellido"],
                fila["puesto"] or "Analista",
                fila["rol"],
                fila["username"],
                fila["rendimiento_futuro"],
                fila["clasificacion_rendimiento"]
            ])

        # Tabla Promedios por Puesto
        ws.append(["", "", "", ""])
        ws.append(["Puesto", "Promedio de Rendimiento Futuro"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for fila in promedios_por_puesto:
            ws.append([fila["puesto"], fila["promedio"]])

        # Guardar gráficos dinámicos
        ruta_grafico1 = os.path.join(TEMP_DIR, "grafico_ranking.png")
        ruta_grafico2 = os.path.join(TEMP_DIR, "grafico_puesto.png")
        with open(ruta_grafico1, "wb") as f:
            f.write(base64.b64decode(grafico_base64))
        with open(ruta_grafico2, "wb") as f:
            f.write(base64.b64decode(grafico_puesto_base64))

        img1 = ExcelImage(ruta_grafico1)
        img2 = ExcelImage(ruta_grafico2)
        img1.width = img2.width = 600
        img1.height = img2.height = 300
        ws.add_image(img1, f"A{ws.max_row + 2}")
        ws.add_image(img2, f"A{ws.max_row + 20}")

        # Insertar imágenes externas del frontend
        imagenes_custom = [
            "cantidad_analistas.png",
            "distribucion_analistas.png",
            "promedios_analistas.png",
            "rendimiento_analistas.png"
        ]
        imagenes_path = os.path.join(os.getcwd(), "imagenes_reportes")
        fila_inicio = ws.max_row + 2
        columna_inicio = "G"

        for i, nombre in enumerate(imagenes_custom):
            ruta_imagen = os.path.join(imagenes_path, nombre)
            if os.path.exists(ruta_imagen):
                try:
                    img = ExcelImage(ruta_imagen)
                    img.width = 600
                    img.height = 300
                    celda = f"{columna_inicio}{fila_inicio + i * 18}"
                    ws.add_image(img, celda)
                except Exception as e:
                    print(f"Error insertando imagen {nombre}:", e)

        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        archivo_excel = f"desempeno_futuro_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_excel = os.path.join(TEMP_DIR, archivo_excel)
        wb.save(ruta_excel)

        return send_file(ruta_excel, as_attachment=True, download_name=archivo_excel)

    return {"error": "Formato no soportado"}, 400




@reportes_bp.route("/reportes-licencias", methods=["GET"])
@role_required(["manager", "reclutador"])
def reporte_licencias_manager():
    formato = request.args.get("formato", "pdf")
    ids_str = request.args.get("ids")
    ids_filtrados = list(map(int, ids_str.split(","))) if ids_str else None
    
    if not ids_filtrados:
        return {"error": "Debe especificar al menos un ID de licencia para generar el reporte."}, 400

    
    licencias_filtradas = Licencia.query.filter(Licencia.id.in_(ids_filtrados)).all() if ids_filtrados else []

    id_manager = get_jwt_identity()
    manager = Usuario.query.get(id_manager)
    preferencia = Preferencias_empresa.query.get(manager.id_empresa)

    empresa = Empresa.query.get(manager.id_empresa)

    dias_por_tipo = defaultdict(int)
    ranking_empleados = defaultdict(int)
    frecuencia_empleado = defaultdict(int)
    dias_por_mes = defaultdict(int)

    for lic in licencias_filtradas:
        if lic.fecha_inicio and lic.fecha_fin:
            dias = (lic.fecha_fin - lic.fecha_inicio).days
            dias = max(dias, 0)

            dias_por_tipo[lic.tipo] += dias
            ranking_empleados[lic.id_empleado] += dias
            frecuencia_empleado[lic.id_empleado] += 1
            dias_por_mes[lic.fecha_inicio.month] += dias

    
    dias_por_tipo = [{"tipo": k, "total_dias": v} for k, v in dias_por_tipo.items()]
    ranking_empleados = [{"username": Usuario.query.get(id_).username, "total_dias": v} for id_, v in sorted(ranking_empleados.items(), key=lambda x: -x[1])[:10]]
    frecuencia_empleado = [{"username": Usuario.query.get(id_).username, "cantidad_licencias": v} for id_, v in frecuencia_empleado.items()]
    dias_mes_data = [{"mes": ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"][k - 1], "total_dias": v} for k, v in sorted(dias_por_mes.items())]

    grafico_base64 = grafico_ausencias_base64(dias_mes_data)
    
    if formato == "pdf":
        env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), '..', 'templates')))
        template = env.get_template("reporte_asistencia_profesional.html")

        html_out = template.render(
            empresa=empresa.nombre,
            logo_url=preferencia.logo_url if preferencia and preferencia.logo_url else None,
            color=preferencia.color_principal if preferencia and preferencia.color_principal else "#2E86C1",
            dias_por_tipo=dias_por_tipo,
            ranking_empleados=ranking_empleados,
            frecuencia_empleado=frecuencia_empleado,
            grafico_ausencias_base64=grafico_base64,
            now=datetime.now()
        )

        archivo = f"reporte_asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta = os.path.join(TEMP_DIR, archivo)
        os.makedirs(TEMP_DIR, exist_ok=True)
        HTML(string=html_out).write_pdf(ruta)

        return send_file(ruta, as_attachment=True, download_name=archivo)

    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Licencias"

        color = (preferencia.color_principal if preferencia and preferencia.color_principal else "#2E86C1").lstrip("#")        
        logo_path = None
        fila = 1  # control de fila actual

        ws.merge_cells("A1:D1")
        ws["A1"] = f"Reporte de Asistencia - {empresa.nombre}"
        ws["A1"].font = Font(bold=True, size=14, color="000000")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 2

        if preferencia and preferencia.logo_url:
            try:
                response = requests.get(preferencia.logo_url)
                if response.status_code == 200:
                    logo_path = os.path.join(TEMP_DIR, "logo_empresa.png")
                    with open(logo_path, "wb") as f:
                        f.write(response.content)
                    logo = ExcelImage(logo_path)
                    logo.width = 120
                    logo.height = 60
                    ws.add_image(logo, "E1")
            except Exception as e:
                print("No se pudo insertar el logo:", e)

        ws[f"A{fila}"] = "Días por Tipo de Licencia"
        ws[f"A{fila}"].font = Font(bold=True, color="000000")
        ws[f"A{fila}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        ws.append(["Tipo de Licencia", "Total de Días"])
        for cell in ws[f"A{fila}":f"B{fila}"][0]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        for fila_data in dias_por_tipo:
            ws.append([fila_data["tipo"], fila_data["total_dias"]])
            fila += 1

        ws[f"A{fila}"] = "Ranking de Empleados"
        ws[f"A{fila}"].font = Font(bold=True, color="000000")
        ws[f"A{fila}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        ws.append(["#", "Usuario", "Días Totales"])
        for cell in ws[f"A{fila}":f"C{fila}"][0]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        for i, fila_data in enumerate(ranking_empleados, start=1):
            ws.append([i, fila_data["username"], fila_data["total_dias"]])
            fila += 1

        ws[f"A{fila}"] = "Frecuencia por Usuario"
        ws[f"A{fila}"].font = Font(bold=True, color="000000")
        ws[f"A{fila}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        ws.append(["Usuario", "Cantidad de Licencias"])
        for cell in ws[f"A{fila}":f"B{fila}"][0]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        for fila_data in frecuencia_empleado:
            ws.append([fila_data["username"], fila_data["cantidad_licencias"]])
            fila += 1

        ws[f"A{fila}"] = "Ausencias por Mes"
        ws[f"A{fila}"].font = Font(bold=True, color="000000")
        ws[f"A{fila}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        ws.append(["Mes", "Total de Días"])
        for cell in ws[f"A{fila}":f"B{fila}"][0]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        fila += 1
        for fila_data in dias_mes_data:
            ws.append([fila_data["mes"], fila_data["total_dias"]])
            fila += 1

        ruta_grafico = os.path.join(TEMP_DIR, "grafico_ausencias.png")
        with open(ruta_grafico, "wb") as f:
            f.write(base64.b64decode(grafico_base64))

        try:
            img_chart = ExcelImage(ruta_grafico)
            img_chart.width = 600
            img_chart.height = 300
            ws.add_image(img_chart, f"A{fila + 2}")
        except Exception as e:
            print("No se pudo insertar el gráfico:", e)

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_len + 2

        archivo_excel = f"reporte_asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta_excel = os.path.join(TEMP_DIR, archivo_excel)
        wb.save(ruta_excel)

    return send_file(ruta_excel, as_attachment=True, download_name=archivo_excel)


@reportes_bp.route("/reportes-riesgos", methods=["GET"])
@role_required(["manager", "reclutador"])
def reporte_riesgos():
    formato = request.args.get("formato", "pdf")

    id_manager = get_jwt_identity()
    manager = Usuario.query.get(id_manager)
    empresa = Empresa.query.get(manager.id_empresa)
    preferencia = Preferencias_empresa.query.get(manager.id_empresa)

    color_excel = (preferencia.color_principal if preferencia and preferencia.color_principal else "#2E86C1").lstrip("#")
    color_mpl = "#" + color_excel    

    riesgos = db.session.query(
        Usuario.username,
        Usuario.puesto_trabajo,
        Rol.nombre.label("rol"),
        RendimientoEmpleado.riesgo_despido_predicho,
        RendimientoEmpleado.riesgo_renuncia_predicho,
        RendimientoEmpleado.riesgo_rotacion_predicho
    ).join(RendimientoEmpleado, Usuario.id == RendimientoEmpleado.id_usuario)\
     .join(UsuarioRol, Usuario.id == UsuarioRol.id_usuario)\
     .join(Rol, UsuarioRol.id_rol == Rol.id)\
     .filter(Usuario.id_empresa == empresa.id)\
     .filter(RendimientoEmpleado.riesgo_despido_predicho != None).all()

    resumen_despido = {}
    resumen_renuncia = {}
    resumen_rotacion = {}
    tabla_completa = []

    def agrupar(dic, clave):
        dic[clave] = dic.get(clave, 0) + 1

    for r in riesgos:
        tabla_completa.append({
            "username": r.username,
            "puesto": r.puesto_trabajo,
            "rol": r.rol,
            "riesgo_despido": r.riesgo_despido_predicho,
            "riesgo_renuncia": r.riesgo_renuncia_predicho,
            "riesgo_rotacion": r.riesgo_rotacion_predicho
        })
        agrupar(resumen_despido, r.riesgo_despido_predicho)
        agrupar(resumen_renuncia, r.riesgo_renuncia_predicho)
        agrupar(resumen_rotacion, r.riesgo_rotacion_predicho)

    def generar_grafico(dic, titulo, nombre_archivo):
        fig, ax = plt.subplots()
        ax.bar(dic.keys(), dic.values(), color=color_mpl)
        ax.set_title(titulo)
        img = BytesIO()
        plt.tight_layout()
        plt.savefig(img, format='png')
        plt.close(fig)
        img.seek(0)
        ruta = os.path.join(TEMP_DIR, nombre_archivo)
        with open(ruta, 'wb') as f:
            f.write(img.getvalue())
        return ruta

    os.makedirs(TEMP_DIR, exist_ok=True)
    path_despido = generar_grafico(resumen_despido, "Riesgo de Despido", "grafico_despido.png")
    path_renuncia = generar_grafico(resumen_renuncia, "Riesgo de Renuncia", "grafico_renuncia.png")
    path_rotacion = generar_grafico(resumen_rotacion, "Riesgo de Rotación", "grafico_rotacion.png")

    def img_to_base64(path):
        with open(path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode("utf-8")

    if formato == "pdf":
        env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), '..', 'templates')))
        template = env.get_template("reporte_riesgos.html")

        ruta_img_despido = os.path.join(BASE_DIR, 'imagenes_reportes', 'riesgo_despido_analistas.png')
        imagen_despido = imagen_a_base64(ruta_img_despido)

        ruta_img_rotacion = os.path.join(BASE_DIR, 'imagenes_reportes', 'riesgo_rotacion_analistas.png')
        imagen_rotacion = imagen_a_base64(ruta_img_rotacion)

        ruta_img_renuncia = os.path.join(BASE_DIR, 'imagenes_reportes', 'riesgo_renuncia_analistas.png')
        imagen_renuncia = imagen_a_base64(ruta_img_renuncia)

        html_out = template.render(
            empresa=empresa.nombre,
            logo_url=preferencia.logo_url if preferencia and preferencia.logo_url else None,
            color="#" + color_excel,
            tabla_completa=tabla_completa,
            resumen_despido=resumen_despido,
            resumen_renuncia=resumen_renuncia,
            resumen_rotacion=resumen_rotacion,
            grafico_despido_base64=img_to_base64(path_despido),
            grafico_renuncia_base64=img_to_base64(path_renuncia),
            grafico_rotacion_base64=img_to_base64(path_rotacion),
            imagen_despido=imagen_despido,
            imagen_rotacion=imagen_rotacion,
            imagen_renuncia=imagen_renuncia,
            now=datetime.now()
        )

        archivo = f"reporte_riesgos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        ruta = os.path.join(TEMP_DIR, archivo)
        HTML(string=html_out).write_pdf(ruta)
        return send_file(ruta, as_attachment=True, download_name=archivo)

    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Riesgos"

        ws.merge_cells("A1:F1")
        ws["A1"] = f"Reporte de Riesgos - {empresa.nombre}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        if preferencia and preferencia.logo_url:
            try:
                response = requests.get(preferencia.logo_url)
                img = Image.open(BytesIO(response.content))
                logo_path = os.path.join(TEMP_DIR, "logo_empresa.png")
                img.save(logo_path)
                logo = ExcelImage(logo_path)
                logo.width = 110
                logo.height = 45
                ws.add_image(logo, "A1")
            except:
                pass

        ws.append([""])
        ws.append(["Usuario", "Puesto", "Rol", "Riesgo Despido", "Riesgo Renuncia", "Riesgo Rotación"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=color_excel, end_color=color_excel, fill_type="solid")

        for fila in tabla_completa:
            ws.append([
                fila["username"],
                fila["puesto"] or "Analista",
                fila["rol"],
                fila["riesgo_despido"],
                fila["riesgo_renuncia"],
                fila["riesgo_rotacion"]
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

        for i, path in enumerate([path_despido, path_renuncia, path_rotacion]):
            img = ExcelImage(path)
            img.width = 500
            img.height = 300
            ws.add_image(img, f"A{ws.max_row + 3 + i*15}")

        archivo = f"reporte_riesgos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ruta = os.path.join(TEMP_DIR, archivo)
        wb.save(ruta)
        return send_file(ruta, as_attachment=True, download_name=archivo)

    return {"error": "Formato no soportado"}, 400




def imagen_a_base64(ruta_local):
    with open(ruta_local, "rb") as img_file:
        encoded_string = base64.b64encode(img_file.read()).decode("utf-8")
        return f"data:image/png;base64,{encoded_string}"
    

@reportes_bp.route("/reporte-eficacia-reclutadores", methods=["GET"])
@role_required(["manager"])
def reporte_eficacia_reclutadores():
    formato = request.args.get("formato", "pdf")
    ids_str = request.args.get("ids")
    print(ids_str)
    ids_filtrados = list(map(int, ids_str.split(","))) if ids_str else None
    
    id_manager = get_jwt_identity()
    manager = Usuario.query.get(id_manager)
    empresa = Empresa.query.get(manager.id_empresa)
    preferencia = Preferencias_empresa.query.get(manager.id_empresa)

    color_hex = (getattr(preferencia, "color_principal", None) or "#2E86C1").lstrip("#")
    color_mpl = f"#{color_hex}"

    sub_reclutadores = (
        db.session.query(
            Usuario.id.label("id_reclutador"),
            Usuario.username.label("reclutador")
        )
        .join(Oferta_analista, Oferta_analista.id_analista == Usuario.id)
        .filter(Oferta_analista.id_oferta.in_(ids_filtrados))
        .filter(Usuario.id_empresa == empresa.id)
        .distinct()
        .subquery()
    )

    datos = (
        db.session.query(
            sub_reclutadores.c.reclutador,
            func.count(Job_Application.id).label("postulaciones"),
            func.sum(case((Job_Application.is_apto == True, 1), else_=0)).label("contratados")
        )
        .join(Oferta_analista, Oferta_analista.id_analista == sub_reclutadores.c.id_reclutador)
        .outerjoin(Job_Application, Job_Application.id_oferta == Oferta_analista.id_oferta)
        .filter(Oferta_analista.id_oferta.in_(ids_filtrados))
        .group_by(sub_reclutadores.c.reclutador)
        .all()
    )


    tabla = []
    for r in datos:
        tasa = round((r.contratados / r.postulaciones) * 100, 2) if r.postulaciones else 0
        tabla.append({
            "reclutador": r.reclutador,
            "postulaciones": r.postulaciones,
            "contratados": r.contratados,
            "tasa_conversion": tasa
        })

    logo_url = preferencia.logo_url if preferencia and getattr(preferencia, "logo_url", None) else None

    if formato == "excel":
        return generar_excel_eficacia(tabla, empresa.nombre, logo_url, color_hex)
    elif formato == "pdf":
        return generar_pdf_eficacia(tabla, empresa.nombre, logo_url, f"#{color_hex}")
    else:
        return {"error": "Formato no soportado"}, 400

def generar_pdf_eficacia(tabla, nombre_empresa, logo_url, color):
    env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__), '..', 'templates')))
    template = env.get_template("reporte_eficacia_reclutadores.html")

    etiquetas = [r["reclutador"] for r in tabla]
    valores = [r["tasa_conversion"] for r in tabla]

    fig, ax = plt.subplots()
    ax.bar(etiquetas, valores, color=color)
    ax.set_title("Tasa de Conversión por Reclutador")
    ax.set_ylabel("Tasa (%)")
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    img = BytesIO()
    plt.savefig(img, format='png', dpi=150)
    plt.close(fig)
    img.seek(0)
    grafico_base64 = base64.b64encode(img.getvalue()).decode("utf-8")

    etiquetas = [fila["reclutador"] for fila in tabla if fila["contratados"] > 0]
    valores = [fila["contratados"] for fila in tabla if fila["contratados"] > 0]

    fig, ax = plt.subplots()
    ax.pie(valores, labels=etiquetas, autopct='%1.1f%%', startangle=90)
    ax.axis('equal')  # Para que sea circular

    img = BytesIO()
    plt.tight_layout()
    plt.savefig(img, format='png')
    plt.close(fig)
    img.seek(0)

    grafico_torta_base64 = base64.b64encode(img.read()).decode("utf-8")

    html_renderizado = template.render(
        empresa=nombre_empresa,
        logo_url=logo_url,
        color=color,
        tabla=tabla,
        fecha=datetime.now().strftime('%d/%m/%Y %H:%M'),
        grafico_base64=grafico_base64,
        grafico_torta_base64=grafico_torta_base64,
    )

    nombre_archivo = f"eficacia_reclutadores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    path = os.path.join(TEMP_DIR, nombre_archivo)

    HTML(string=html_renderizado).write_pdf(path)
    return send_file(path, as_attachment=True, download_name=nombre_archivo)


def generar_excel_eficacia(tabla, nombre_empresa, logo_url, color_hex):
    wb = Workbook()
    ws = wb.active
    ws.title = "Eficacia Reclutadores"

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Reporte de Eficacia - {nombre_empresa}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    if logo_url:
        try:
            response = requests.get(logo_url)
            img = Image.open(BytesIO(response.content))
            logo_path = os.path.join("temp", "logo_temp.png")
            img.save(logo_path)
            excel_img = ExcelImage(logo_path)
            excel_img.width = 110
            excel_img.height = 45
            ws.add_image(excel_img, "A1")
        except:
            pass

    ws.append([""])
    ws.append(["Reclutador", "Postulaciones", "Contratados", "Tasa de Conversión (%)"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

    for fila in tabla:
        ws.append([
            fila["reclutador"],
            fila["postulaciones"],
            fila["contratados"],
            fila["tasa_conversion"]
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    fig, ax = plt.subplots()
    ax.bar(
        [r["reclutador"] for r in tabla],
        [r["tasa_conversion"] for r in tabla],
        color=f"#{color_hex}"
    )
    ax.set_title("Tasa de Conversión por Reclutador")
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    ruta_temp = os.path.join("temp", "grafico_excel.png")
    os.makedirs("temp", exist_ok=True)
    plt.savefig(ruta_temp)
    plt.close()

    ultima_fila = ws.max_row + 2

    img_excel = ExcelImage(ruta_temp)
    img_excel.width = 500
    img_excel.height = 300
    ws.add_image(img_excel, f"A{ultima_fila}")

    etiquetas = [fila["reclutador"] for fila in tabla if fila["contratados"] > 0]
    valores = [fila["contratados"] for fila in tabla if fila["contratados"] > 0]

    fig, ax = plt.subplots()
    ax.pie(valores, labels=etiquetas, autopct='%1.1f%%', startangle=90)
    ax.axis('equal')  # Asegura forma circular
    plt.tight_layout()

    path = os.path.join(TEMP_DIR, "grafico_torta_contratados.png")
    plt.savefig(path, format='png')
    plt.close(fig)

    path_grafico_torta = path

    img_torta = ExcelImage(path_grafico_torta)
    img_torta.width = 500
    img_torta.height = 300
    ws.add_image(img_torta, f"A{ultima_fila + 20}")

    nombre_archivo = f"eficacia_reclutadores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(TEMP_DIR, nombre_archivo)
    wb.save(path)
    return send_file(path, as_attachment=True, download_name=nombre_archivo)



